"""
StaffKit Debug Helper
Usage:
    python scripts/db_query.py                      # show summary
    python scripts/db_query.py "SELECT ..."         # run custom SQL
    python scripts/db_query.py --groups             # show group counts
    python scripts/db_query.py --compare <xlsx>    # compare Excel vs DB
"""
import sqlite3
import sys
import os

DB_PATH = r"C:\Users\Zonzon\AppData\Local\io.staffkit.app\staff_kit.sqlite3"


def get_conn():
    return sqlite3.connect(f"file:{DB_PATH}?mode=ro", uri=True)


def show_summary(conn):
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM employees")
    total = cur.fetchone()[0]
    print(f"Total employees: {total}")

    cur.execute("""
        SELECT COALESCE(NULLIF(TRIM(staff_group),''), 'employee_list') AS grp, COUNT(*)
        FROM employees GROUP BY grp ORDER BY grp
    """)
    print("\nBy group:")
    for grp, cnt in cur.fetchall():
        print(f"  {grp:25s} {cnt}")


def show_groups(conn):
    show_summary(conn)


def run_sql(conn, sql):
    cur = conn.cursor()
    cur.execute(sql)
    rows = cur.fetchall()
    if cur.description:
        headers = [d[0] for d in cur.description]
        print(" | ".join(headers))
        print("-" * 60)
        for row in rows:
            print(" | ".join(str(v) if v is not None else "NULL" for v in row))
    print(f"\n({len(rows)} rows)")


def compare_excel(conn, xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    header = [str(c.value).strip() if c.value else "" for c in ws[1]]
    print("Excel headers:", header[:10])

    # Normalize header keys the same as Rust
    def norm_key(v):
        return "".join(c.lower() for c in str(v) if c.isalnum())

    headers_map = {norm_key(h): i for i, h in enumerate(header) if h}
    eid_col = next((headers_map[k] for k in ["eeid","emid","employeeid","staffid"] if k in headers_map), None)
    name_col = next((headers_map[k] for k in ["vietnamesename","fullname","name"] if k in headers_map), None)
    email_col = next((headers_map[k] for k in ["workingemail","email"] if k in headers_map), None)

    print(f"Columns detected: eid={eid_col}, name={name_col}, email={email_col}\n")

    excel_eids = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        eid = str(row[eid_col]).strip().upper() if eid_col is not None and row[eid_col] else ""
        if eid:
            excel_eids[eid] = {"row": i, "name": str(row[name_col]).strip() if name_col else "", "email": str(row[email_col]).strip() if email_col else ""}

    cur = conn.cursor()
    cur.execute("SELECT employee_id, full_name, staff_group FROM employees")
    db_map = {row[0].upper(): {"name": row[1], "group": row[2]} for row in cur.fetchall()}

    print(f"Excel: {len(excel_eids)} EE.IDs | DB: {len(db_map)} employees")

    missing = set(excel_eids) - set(db_map)
    extra = set(db_map) - set(excel_eids)
    in_wrong_group = {k: db_map[k] for k in set(excel_eids) & set(db_map) if db_map[k]["group"] != "employee_list"}

    if missing:
        print(f"\n❌ In Excel but NOT in DB ({len(missing)}):")
        for eid in sorted(missing):
            d = excel_eids[eid]
            print(f"  Row {d['row']}: {eid} - {d['name']} ({d['email']})")
    else:
        print("\n✅ All Excel EE.IDs found in DB")

    if in_wrong_group:
        print(f"\n⚠️  In Excel but in WRONG group ({len(in_wrong_group)}):")
        for eid, d in in_wrong_group.items():
            print(f"  {eid}: {d['name']} → group={d['group']}")
    else:
        print("✅ All Excel employees are in employee_list")

    print(f"\nDB-only (not in Excel): {len(extra)} employees")


def main():
    args = sys.argv[1:]
    conn = get_conn()

    if not args:
        show_summary(conn)
    elif args[0] == "--groups":
        show_groups(conn)
    elif args[0] == "--compare" and len(args) > 1:
        compare_excel(conn, args[1])
    else:
        run_sql(conn, " ".join(args))

    conn.close()


if __name__ == "__main__":
    main()
